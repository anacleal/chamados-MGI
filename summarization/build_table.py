import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESUMOS_DIR = "outLLM/detailed_summarization"

K_POR_SISTEMA = {
    "SIASS":  6,
    "SIAPE":  8,
    "SIGEPE": 5,
    "SOUGOV": 5,
    "TOTAIS": 10,
}

def load_file(path):
    if not os.path.exists(path):
        return None
    with open(path, encoding="utf-8") as f:
        return f.read().strip()

def count_docs(sistema, topic):
    """Conta documentos do tópico via Resumo_Topicos_Dominantes.csv."""
    path = f"../topic_modeling/bertopic_resultados/{sistema}/Resumo_Topicos_Dominantes.csv"
    if not os.path.exists(path):
        return None
    df = pd.read_csv(path)
    col = "dominant_topic" if "dominant_topic" in df.columns else df.columns[-1]
    return int((df[col] == topic).sum())

def parse_resumo(texto):
    """Separa Padrão Dominante e Impacto Operacional do resumo."""
    padrao, impacto = "", ""
    if not texto:
        return padrao, impacto
    m = re.search(r'\*\*Padrão Dominante\*\*:\s*(.*?)(?=\*\*Impacto Operacional\*\*|$)', texto, re.S)
    if m:
        padrao = m.group(1).strip()
    m = re.search(r'\*\*Impacto Operacional\*\*:\s*(.*)', texto, re.S)
    if m:
        impacto = m.group(1).strip()
    return padrao, impacto

rows = []
for sis, k in K_POR_SISTEMA.items():
    for t in range(k):
        base = f"{RESUMOS_DIR}/{sis}"
        titulo  = load_file(f"{base}/titulo_topic_{t}.txt") or ""
        resumo  = load_file(f"{base}/summary_topic_{t}.txt") or ""
        padrao, impacto = parse_resumo(resumo)
        n_docs = count_docs(sis, t)
        rows.append({
            "Sistema":            sis,
            "Tópico":             t,
            "Título":             titulo,
            "Nº Documentos":      n_docs,
            "Padrão Dominante":   padrao,
            "Impacto Operacional": impacto,
            "Observações":        "",   # coluna vazia para a especialista anotar
        })

# ── Excel ────────────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Revisão de Tópicos"

HEADERS = ["Sistema", "Tópico", "Título", "Nº Documentos",
           "Padrão Dominante", "Impacto Operacional", "Observações"]

# Cores por sistema
COR_SISTEMA = {
    "SIASS":  "D6E4F0",
    "SIAPE":  "D5F5E3",
    "SIGEPE": "FCF3CF",
    "SOUGOV": "FADBD8",
    "TOTAIS": "E8DAEF",
}

# Cabeçalho
header_fill = PatternFill("solid", start_color="2C3E50")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
thin        = Side(style="thin", color="BDBDBD")
border      = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = center
    cell.border    = border

ws.row_dimensions[1].height = 32

# Dados
for r, row in enumerate(rows, 2):
    cor = COR_SISTEMA.get(row["Sistema"], "FFFFFF")
    fill = PatternFill("solid", start_color=cor)

    for col, key in enumerate(HEADERS, 1):
        val  = row.get(key, "")
        cell = ws.cell(row=r, column=col, value=val)
        cell.fill   = fill
        cell.border = border
        cell.font   = Font(name="Arial", size=10)
        if col in (1, 2, 4):   # Sistema, Tópico, Nº Docs — centralizado
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = left_wrap

    ws.row_dimensions[r].height = 80

# Larguras das colunas
COL_WIDTHS = {1: 10, 2: 9, 3: 30, 4: 14, 5: 55, 6: 45, 7: 35}
for col, w in COL_WIDTHS.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# Congela cabeçalho
ws.freeze_panes = "A2"

# Auto-filtro
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

out = os.path.join(os.path.dirname(__file__), "revisao_topicos.xlsx")
wb.save(out)
print(f"Salvo em {out}")